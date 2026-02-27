"""
Экспорт данных кафе в Excel
"""

from datetime import datetime
from typing import Optional
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from database import CafeDatabase


class ExcelExporter:
    """Класс для экспорта данных в Excel"""

    def __init__(self, db: CafeDatabase):
        self.db = db

    def export_menu_with_costs(self, filename: str = "cafe_menu.xlsx"):
        """Экспорт меню с себестоимостью и наценкой"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Меню и себестоимость"

        # Заголовки
        headers = ["ID", "Название блюда", "Категория", "Цена продажи",
                  "Себестоимость", "Прибыль с порции", "Наценка %", "Маржа %"]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные
        dishes = self.db.get_dishes()
        for row, dish in enumerate(dishes, 2):
            margin_info = self.db.get_dish_margin(dish.id)

            ws.cell(row=row, column=1, value=dish.id)
            ws.cell(row=row, column=2, value=dish.name)
            ws.cell(row=row, column=3, value=dish.category)
            ws.cell(row=row, column=4, value=margin_info.get('price', dish.price))
            ws.cell(row=row, column=5, value=margin_info.get('cost', 0))
            ws.cell(row=row, column=6, value=margin_info.get('margin_amount', 0))
            ws.cell(row=row, column=7, value=margin_info.get('markup_percent', 0))
            ws.cell(row=row, column=8, value=margin_info.get('margin_percent', 0))

        # Автоширина столбцов
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(filename)
        return filename

    def export_tech_cards(self, filename: str = "tech_cards.xlsx"):
        """Экспорт всех техкарт"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Техкарты"

        dishes = self.db.get_dishes()

        row = 1
        for dish in dishes:
            # Заголовок блюда
            header_cell = ws.cell(row=row, column=1, value=f"Блюдо: {dish.name}")
            header_cell.font = Font(bold=True, size=14)
            header_cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

            row += 1
            # Заголовки таблицы техкарты
            tech_headers = ["Ингредиент", "Единица измерения", "Количество", "Цена за единицу", "Сумма"]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(tech_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            row += 1

            # Ингредиенты
            recipe = self.db.get_recipe(dish.id)
            total_cost = 0
            for item in recipe:
                ingredient = self.db.get_ingredient(item.ingredient_id)
                if ingredient:
                    item_cost = ingredient.price_per_unit * item.quantity
                    total_cost += item_cost

                    ws.cell(row=row, column=1, value=ingredient.name)
                    ws.cell(row=row, column=2, value=ingredient.unit)
                    ws.cell(row=row, column=3, value=item.quantity)
                    ws.cell(row=row, column=4, value=ingredient.price_per_unit)
                    ws.cell(row=row, column=5, value=item_cost)
                    row += 1

            # Итого себестоимость
            total_cell = ws.cell(row=row, column=4, value="Итого себестоимость:")
            total_cell.font = Font(bold=True)
            ws.cell(row=row, column=5, value=total_cost).font = Font(bold=True)
            row += 2

        # Автоширина столбцов
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(filename)
        return filename

    def export_financial_report(self, start_date: Optional[str] = None,
                               end_date: Optional[str] = None,
                               filename: str = "financial_report.xlsx"):
        """Экспорт финансового отчета"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Лист 1: Общая финансовая сводка
        ws_summary = wb.active
        ws_summary.title = "Сводка"

        profit_data = self.db.get_profit(start_date, end_date)

        ws_summary.cell(row=1, column=1, value="Финансовый отчет").font = Font(bold=True, size=16)
        if start_date and end_date:
            ws_summary.cell(row=2, column=1, value=f"Период: {start_date} - {end_date}")
        else:
            ws_summary.cell(row=2, column=1, value="За весь период")

        row = 4
        summary_data = [
            ("Выручка", profit_data['revenue']),
            ("Себестоимость проданных товаров", profit_data['cost_of_goods_sold']),
            ("Валовая прибыль", profit_data['gross_profit']),
            ("Расходы", profit_data['total_expenses']),
            ("Чистая прибыль", profit_data['net_profit'])
        ]

        for label, value in summary_data:
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1

        # Лист 2: Расходы по категориям
        ws_expenses = wb.create_sheet("Расходы по категориям")
        expenses_by_category = self.db.get_expenses_by_category(start_date, end_date)

        ws_expenses.cell(row=1, column=1, value="Категория").font = Font(bold=True)
        ws_expenses.cell(row=1, column=2, value="Сумма").font = Font(bold=True)

        row = 2
        for category, amount in sorted(expenses_by_category.items(), key=lambda x: x[1], reverse=True):
            ws_expenses.cell(row=row, column=1, value=category)
            ws_expenses.cell(row=row, column=2, value=amount)
            row += 1

        # Лист 3: Детальные расходы
        ws_expenses_detailed = wb.create_sheet("Детальные расходы")
        expenses = self.db.get_expenses(start_date, end_date)

        headers = ["Дата", "Категория", "Сумма", "Описание"]
        for col, header in enumerate(headers, 1):
            cell = ws_expenses_detailed.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row, expense in enumerate(expenses, 2):
            ws_expenses_detailed.cell(row=row, column=1, value=expense.date)
            ws_expenses_detailed.cell(row=row, column=2, value=expense.category)
            ws_expenses_detailed.cell(row=row, column=3, value=expense.amount)
            ws_expenses_detailed.cell(row=row, column=4, value=expense.description or "")

        # Лист 4: Продажи
        ws_sales = wb.create_sheet("Продажи")
        sales = self.db.get_sales(start_date, end_date)

        headers = ["Дата", "Блюдо", "Количество", "Сумма"]
        for col, header in enumerate(headers, 1):
            cell = ws_sales.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row, sale in enumerate(sales, 2):
            ws_sales.cell(row=row, column=1, value=sale.date)
            ws_sales.cell(row=row, column=2, value=sale.dish_name)
            ws_sales.cell(row=row, column=3, value=sale.quantity)
            ws_sales.cell(row=row, column=4, value=sale.total_amount)

        # Автоширина столбцов
        for ws in wb.worksheets:
            for col in ws.iter_cols():
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        wb.save(filename)
        return filename

